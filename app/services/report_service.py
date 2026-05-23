"""Report generation (SPEC §6) — CSV and XLSX now, PDF planned next.

Assembles the sections the frontend's report builder offered (summary,
holdings, allocation, performance, risk, transactions, valuation,
Monte-Carlo) from the analytics layer and renders them. Everything is
computed on the backend; the endpoint streams a downloadable file.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime, timezone

import numpy as np

from app.config import settings
from app.models import Asset, User
from app.schemas.report import ReportFormat, ReportSection
from app.services import analytics_service, transaction_service
from app.services.analytics import metrics as metrics_mod
from app.services.analytics import montecarlo


@dataclass
class Section:
    title: str
    headers: list[str]
    rows: list[list]


@dataclass
class Report:
    filename: str
    media_type: str
    content: bytes


def _num(x) -> float:
    return float(x) if x is not None else 0.0


async def _gather(user: User, portfolio_id: int | None, days: int | None):
    """Resolve the shared datasets (records + equity series) once."""
    records = await analytics_service.valued_positions(user, portfolio_id)
    if portfolio_id is None:
        dash = await analytics_service.dashboard(user, days=days)
        series = dash["series"]
    else:
        series = await analytics_service.portfolio_series(
            user, portfolio_id, days=days
        )
    return records, series


async def _build_sections(
    user: User,
    portfolio_id: int | None,
    sections: list[ReportSection],
    days: int | None,
) -> list[Section]:
    records, series = await _gather(user, portfolio_id, days)
    values = np.array(series.get("values", []), dtype=float)
    out: list[Section] = []
    wanted = set(sections)

    total = sum(float(r["market_value"]) for r in records)
    cost = sum(float(r["cost"]) for r in records)
    pnl = total - cost

    if ReportSection.SUMMARY in wanted:
        rows = [
            ["Стоимость портфеля", round(total, 2), settings.base_currency],
            ["Стоимость покупок", round(cost, 2), settings.base_currency],
            ["P&L (нереализ.)", round(pnl, 2), settings.base_currency],
            ["P&L %", round((pnl / cost * 100) if cost else 0.0, 2), "%"],
            ["Позиций", len(records), ""],
        ]
        out.append(Section("Сводка", ["Показатель", "Значение", "Ед."], rows))

    if ReportSection.HOLDINGS in wanted:
        rows = [
            [
                r["symbol"],
                r["name"],
                r["asset_class"],
                float(r["quantity"]),
                round(float(r["current_price"]), 4),
                round(float(r["market_value"]), 2),
                round(float(r["pnl"]), 2),
                round(r["pnl_pct"] * 100, 2),
            ]
            for r in sorted(records, key=lambda r: float(r["market_value"]), reverse=True)
        ]
        out.append(
            Section(
                "Состав портфеля",
                ["Тикер", "Название", "Класс", "Кол-во", "Цена", "Стоимость", "P&L", "P&L %"],
                rows,
            )
        )

    if ReportSection.ALLOCATION in wanted:
        from app.services.analytics import concentration

        rows = [
            [r["label"], round(r["value"], 2), round(r["share"] * 100, 2)]
            for r in concentration.breakdown(records, "asset_class")
        ]
        out.append(Section("Распределение по классам", ["Класс", "Стоимость", "Доля %"], rows))

    if ReportSection.PERFORMANCE in wanted:
        rows = [[d, round(v, 2)] for d, v in zip(series.get("dates", []), values.tolist())]
        out.append(Section("Динамика стоимости", ["Дата", "Стоимость"], rows))

    if ReportSection.RISK in wanted and values.size >= 2:
        m = metrics_mod.compute_all(values, rf=settings.risk_free_rate)
        rows = [
            ["Годовая доходность", round(m["annual_return"] * 100, 2), "%"],
            ["Волатильность (σ)", round(m["volatility"] * 100, 2), "%"],
            ["Sharpe", round(m["sharpe"], 2), ""],
            ["Sortino", round(m["sortino"], 2), ""],
            ["Max Drawdown", round(m["max_drawdown"] * 100, 2), "%"],
            ["Calmar", round(m["calmar"], 2), ""],
            ["VaR 95% (дневной)", round(m["var_95"] * 100, 2), "%"],
            ["CVaR 95%", round(m["cvar_95"] * 100, 2), "%"],
        ]
        out.append(Section("Риск-метрики", ["Метрика", "Значение", "Ед."], rows))

    if ReportSection.TRANSACTIONS in wanted:
        txs = await transaction_service.list_for_user(user, portfolio_id=portfolio_id)
        asset_ids = {t.asset_id for t in txs}
        assets = {a.id: a for a in await Asset.filter(id__in=asset_ids)}
        rows = [
            [
                t.timestamp.date().isoformat(),
                t.tx_type.value,
                assets.get(t.asset_id).symbol if assets.get(t.asset_id) else t.asset_id,
                float(t.quantity),
                float(t.price),
                t.currency,
            ]
            for t in txs
        ]
        out.append(
            Section(
                "Сделки",
                ["Дата", "Тип", "Актив", "Кол-во", "Цена", "Валюта"],
                rows,
            )
        )

    if ReportSection.VALUATION in wanted:
        rows = []
        for aid in {r["asset_id"] for r in records}:
            v = await analytics_service.asset_valuation(aid)
            if not v["has_inputs"]:
                continue
            rows.append(
                [
                    v["symbol"],
                    round(_num(v["current_price"]), 2),
                    round(_num(v["dcf"]["value"]), 2),
                    round(_num(v["gordon"]["value"]), 2),
                    round(_num(v["capm_expected_return"]) * 100, 2),
                ]
            )
        out.append(
            Section(
                "Фундаментальная оценка",
                ["Тикер", "Цена", "DCF", "Gordon", "CAPM E(R) %"],
                rows,
            )
        )

    if ReportSection.MC in wanted and values.size >= 2:
        from app.services.analytics.core import (
            annualized_return,
            annualized_volatility,
            simple_returns,
        )

        mu = annualized_return(values)
        sigma = annualized_volatility(simple_returns(values))
        mc = montecarlo.simulate(float(values[-1]), mu, sigma, n_sims=500)
        rows = [
            ["Старт", round(mc["start_value"], 2)],
            ["Медиана (p50)", round(mc["final"]["p50"], 2)],
            ["Пессимистичный (p10)", round(mc["final"]["p10"], 2)],
            ["Оптимистичный (p90)", round(mc["final"]["p90"], 2)],
            ["P(достичь цели)", round(mc["prob_target"] * 100, 2)],
            ["VaR 5%", round(mc["var_5"] * 100, 2)],
        ]
        out.append(Section("Монте-Карло (1 год)", ["Показатель", "Значение"], rows))

    return out


def _render_csv(sections: list[Section]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    for sec in sections:
        writer.writerow([f"# {sec.title}"])
        writer.writerow(sec.headers)
        writer.writerows(sec.rows)
        writer.writerow([])
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel reads UTF-8


def _render_xlsx(sections: list[Section]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for sec in sections:
        # Sheet titles are <=31 chars and unique.
        title = sec.title[:28] or "Section"
        name, i = title, 1
        while name in used:
            name = f"{title[:26]}_{i}"
            i += 1
        used.add(name)
        ws = wb.create_sheet(name)
        ws.append(sec.headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in sec.rows:
            ws.append(row)
    if not wb.sheetnames:
        wb.create_sheet("Пусто")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


async def generate(user: User, *, portfolio_id, sections, fmt, days=None) -> Report:
    built = await _build_sections(user, portfolio_id, sections, days)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    scope = "all" if portfolio_id is None else f"pf{portfolio_id}"
    if fmt == ReportFormat.CSV:
        return Report(
            filename=f"curs_report_{scope}_{stamp}.csv",
            media_type="text/csv; charset=utf-8",
            content=_render_csv(built),
        )
    return Report(
        filename=f"curs_report_{scope}_{stamp}.xlsx",
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        content=_render_xlsx(built),
    )
